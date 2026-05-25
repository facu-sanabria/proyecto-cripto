# ══════════════════════════════════════════════════════════════════════════════
# excel_writer.py — El "diseñador gráfico" del bot
#
# Este módulo toma los resultados del análisis y los convierte en un Excel
# visualmente claro: colores semáforo, formato financiero y datos ordenados.
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime


# ─── Paleta de colores ─────────────────────────────────────────────────────────
COLORS = {
    "header_bg":   "1F3864",   # azul oscuro para encabezados
    "header_fg":   "FFFFFF",   # texto blanco en encabezados
    "strong_buy":  "00B050",   # verde oscuro
    "buy":         "92D050",   # verde claro
    "neutral":     "FFD966",   # amarillo
    "sell":        "F4B942",   # naranja
    "strong_sell": "FF4444",   # rojo
    "row_even":    "F2F7FF",   # azul muy claro (filas pares)
    "row_odd":     "FFFFFF",   # blanco (filas impares)
    "positive":    "00693E",   # verde oscuro para números positivos
    "negative":    "CC0000",   # rojo oscuro para números negativos
}

# Mapeo señal → color de fondo
SIGNAL_COLORS = {
    "STRONG BUY":  COLORS["strong_buy"],
    "BUY":         COLORS["buy"],
    "NEUTRAL":     COLORS["neutral"],
    "SELL":        COLORS["sell"],
    "STRONG SELL": COLORS["strong_sell"],
}

# ─── Definición de columnas del dashboard ─────────────────────────────────────
# Formato: (Título encabezado, key del dict de resultado, ancho de columna)
COLUMNS = [
    ("🪙 Crypto",        "symbol",      9),
    ("Nombre",           "name",        14),
    ("💰 Precio USD",    "price",       14),
    ("📈 Cambio 24h",    "change_24h",  12),
    ("⭐ Score TA",      "score",       10),
    ("🚦 Señal",         "signal",      14),
    ("RSI",              "rsi",          8),
    ("MACD Hist.",       "macd_hist",   11),
    ("Tendencia EMA",    "ema_trend",   14),
    ("🛑 Stop Loss",     "stop_loss",   13),
    ("🎯 Take Profit",   "take_profit", 13),
    ("R/R Ratio",        "risk_reward", 10),
    ("📝 Por qué",       "reason",      55),
    ("🕐 Actualizado",   "_timestamp",  20),
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border() -> Border:
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _setup_headers(ws):
    """Escribe la fila de encabezados con estilo."""
    ws.freeze_panes = "A2"  # congelar primera fila al hacer scroll

    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = _fill(COLORS["header_bg"])
        cell.font   = Font(color=COLORS["header_fg"], bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28


def _write_data_row(ws, row_idx: int, result: dict):
    """Escribe una fila de datos con formato y colores."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    bg_color  = COLORS["row_even"] if row_idx % 2 == 0 else COLORS["row_odd"]
    signal    = result.get("signal", "NEUTRAL")

    for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
        value = timestamp if key == "_timestamp" else result.get(key, "")
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border    = _border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── Estilos por tipo de columna ─────────────────────────────────────
        if key == "signal":
            # Semáforo de colores según la señal
            color = SIGNAL_COLORS.get(str(value), COLORS["neutral"])
            cell.fill = _fill(color)
            cell.font = Font(bold=True, size=10,
                             color="FFFFFF" if value in ("STRONG BUY", "STRONG SELL") else "1F1F1F")

        elif key == "change_24h":
            cell.fill = _fill(bg_color)
            if isinstance(value, (int, float)):
                cell.value = f"{value:+.2f}%"
                cell.font  = Font(
                    bold=True,
                    color=COLORS["positive"] if value >= 0 else COLORS["negative"]
                )

        elif key == "score":
            cell.fill = _fill(bg_color)
            if isinstance(value, (int, float)):
                cell.font = Font(
                    bold=True,
                    color=COLORS["positive"] if value >= 0 else COLORS["negative"]
                )

        elif key in ("price", "stop_loss", "take_profit"):
            cell.fill = _fill(bg_color)
            if isinstance(value, float) and value > 0:
                # Formato con símbolo $ y decimales apropiados
                cell.number_format = '$#,##0.00####'

        elif key == "reason":
            cell.fill      = _fill(bg_color)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        else:
            cell.fill = _fill(bg_color)

    ws.row_dimensions[row_idx].height = 20


def _apply_score_gradient(ws, num_rows: int):
    """Aplica gradiente rojo→amarillo→verde en columna Score."""
    score_col_idx = next(
        (i + 1 for i, (_, k, _) in enumerate(COLUMNS) if k == "score"), None
    )
    if not score_col_idx:
        return

    col_letter = get_column_letter(score_col_idx)
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{num_rows + 1}",
        ColorScaleRule(
            start_type='num', start_value=-100, start_color='FF4444',
            mid_type='num',   mid_value=0,       mid_color='FFD966',
            end_type='num',   end_value=100,      end_color='00B050'
        )
    )


def _create_info_sheet(wb, num_cryptos: int):
    """Crea una hoja con la leyenda y notas del bot."""
    ws = wb.create_sheet("ℹ️ Leyenda")

    info_rows = [
        ("BOT DE ANÁLISIS CRIPTO", True, "1F3864", "FFFFFF"),
        (f"Última actualización: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", False, None, None),
        (f"Criptomonedas analizadas: {num_cryptos}", False, None, None),
        ("", False, None, None),
        ("INTERPRETACIÓN DE SEÑALES", True, "1F3864", "FFFFFF"),
        ("🟢 STRONG BUY (Score ≥ 60):  Fuerte señal de compra. Múltiples indicadores alineados.", False, "E8F5E9", None),
        ("🟩 BUY (Score ≥ 25):          Señal de compra moderada. Tendencia positiva.", False, "F1F8E9", None),
        ("🟡 NEUTRAL (-25 a +25):       Sin señal clara. Esperar mejor momento.", False, "FFFDE7", None),
        ("🟠 SELL (Score ≤ -25):        Señal de venta moderada. Riesgo creciente.", False, "FFF3E0", None),
        ("🔴 STRONG SELL (Score ≤ -60): Fuerte señal de venta. Evitar entrar.", False, "FFEBEE", None),
        ("", False, None, None),
        ("INDICADORES USADOS", True, "1F3864", "FFFFFF"),
        ("RSI (< 30 = sobrevendido, > 70 = sobrecomprado)", False, None, None),
        ("MACD Histograma (positivo = alcista, negativo = bajista)", False, None, None),
        ("EMA200 (precio sobre la línea = tendencia alcista)", False, None, None),
        ("Bollinger Bands (precio en banda inferior = posible rebote)", False, None, None),
        ("Volumen (volumen alto confirma la tendencia)", False, None, None),
        ("", False, None, None),
        ("STOP LOSS Y TAKE PROFIT", True, "1F3864", "FFFFFF"),
        ("Stop Loss = precio - (1.5 × ATR)   → dónde cortar pérdidas", False, None, None),
        ("Take Profit = precio + (2.5 × ATR) → objetivo de ganancia", False, None, None),
        ("R/R Ratio > 1.5 es generalmente aceptable para operar", False, None, None),
        ("", False, None, None),
        ("⚠️  AVISO IMPORTANTE", True, "B71C1C", "FFFFFF"),
        ("Este bot es una herramienta de análisis educativo.", False, "FFEBEE", "CC0000"),
        ("NO es asesoramiento financiero. Invertir en cripto", False, "FFEBEE", "CC0000"),
        ("implica riesgo de pérdida total del capital.", False, "FFEBEE", "CC0000"),
    ]

    for i, (text, is_header, bg, fg) in enumerate(info_rows, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if is_header:
            cell.fill = _fill(bg)
            cell.font = Font(color=fg, bold=True, size=12)
        elif bg:
            cell.fill = _fill(bg)
            if fg:
                cell.font = Font(color=fg, bold=True)
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 18

    ws.column_dimensions["A"].width = 75


def update_excel(results: list, excel_path: str):
    """
    Crea/sobreescribe el Excel con los resultados del análisis.

    Los resultados deben venir ORDENADOS por score (mejor primero).

    Args:
        results:    Lista de dicts con datos de cada crypto.
        excel_path: Ruta del archivo Excel a generar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "📊 Señales Cripto"
    ws.sheet_view.showGridLines = False  # quitar grilla gris de Excel

    _setup_headers(ws)

    for i, result in enumerate(results):
        _write_data_row(ws, i + 2, result)

    if results:
        _apply_score_gradient(ws, len(results))

    _create_info_sheet(wb, len(results))

    wb.save(excel_path)
    print(f"  ✅ Excel guardado en: {excel_path}")
